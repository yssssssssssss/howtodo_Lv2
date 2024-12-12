import os
import pandas as pd
from tqdm import tqdm
import matplotlib.pyplot as plt
import openpyxl
from openpyxl.drawing.image import Image as XLImage



# z = '女士春夏下装_from_0501'

# 定义基础路径和CSV文件路径
# base_path = f'D://code//data//Lv2期结论//{z}'
# csv_file_path = f'D://code//data//Lv2期结论//{z}//{z}.csv'
# brand_path = f'D://code//data//Lv2期结论//{z}//男鞋品牌分层.xlsx'  # 品牌分类数据源



# 定义不同的筛选条件列表
# filter_layer_cases = [[1.0, 2.0, 3.0], [4.0, 5.0, 6.0]]


for filter_layers in filter_layer_cases:
    # 读取CSV文件
    df = pd.read_csv(csv_file_path)
    df_brand = pd.read_excel(brand_path)  # 读取品牌分类信息

    # 合并df和df_brand数据
    df = pd.merge(df, df_brand, on='main_brand_code', how='left')

    # 筛选数据
    filtered_df = df[df['最终分层'].isin(filter_layers)]

    def extract_matching_part(img_url):
        if pd.isna(img_url):
            return None
        img_url = img_url.split('?')[0]
        img_url = os.path.splitext(img_url)[0]
        parts = img_url.split('/')
        if len(parts) >= 2:
            return f"{parts[-2]}_{parts[-1]}"
        return None

    filtered_df['matching_part'] = filtered_df['img_url'].apply(extract_matching_part)

    all_stats = {}

    for root, dirs, files in os.walk(base_path):
        if 'grounding_output' in dirs:
            grounding_path = os.path.join(root, 'grounding_output')
            folder_name = os.path.basename(root)
            stats = process_grounding_folder(grounding_path)
            all_stats[folder_name] = stats

    # 创建一个Excel工作簿来保存结果（这里每次循环重新创建工作簿，你也可以根据需要调整）
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "Image Distribution Stats"

    # 写入表头
    headers = ["Folder", "Subfolder", "Count", "UV", "Click UV", "CTR"]
    for col, header in enumerate(headers, start=1):
        sheet.cell(row=1, column=col, value=header)

    row = 2  # 从第二行开始写入数据

    # 输出统计结果
    for folder, stats in all_stats.items():
        print(f"\n统计结果 for {folder}:")
        for subfolder, data in stats.items():
            if data['count'] > 0:
                ctr = data['click_uv'] / data['uv'] if data['uv'] > 0 else 0
                print(f"  {subfolder}: 图片数量 = {data['count']}, UV = {data['uv']}, Click UV = {data['click_uv']}, CTR = {ctr:.4f}")

    # 定义颜色映射
    color_map = {
        'price': '#FF6B6B',  # 柔和的红色
        'txt': '#4ECDC4',    # 青绿色
        'scene': '#7986CB',  # 淡紫色
        'white': '#FFD93D'   # 明亮的黄色
    }

    # 绘制每个grounding_output文件夹的饼图并保存数据
    for folder, stats in all_stats.items():
        counts = [data['count'] for data in stats.values() if data['count'] > 0]
        labels = [subfolder for subfolder, data in stats.items() if data['count'] > 0]
        ctrs = [data['ctr'] for data in stats.values() if data['count'] > 0]
        colors = [color_map.get(label, 'gray') for label in labels]  # 使用颜色映射，如果没有指定则默认为灰色

        if counts:  # 只有当有数据时才绘图
            plt.figure(figsize=(12, 9))
            wedges, texts, autotexts = plt.pie(counts, labels=labels, autopct='%1.1f%%', startangle=90, colors=colors)

            # 添加数量和CTR标签
            for i, (autotext, ctr) in enumerate(zip(autotexts, ctrs)):
                autotext.set_text(f'{autotext.get_text()}\n({counts[i]})\nCTR: {ctr:.4f}')

            # 在饼图名中添加筛选条件的值
            plt.title(f"Image Distribution in {folder} ({', '.join([str(layer) for layer in filter_layers])})")
            plt.axis('equal')

            # 添加图例
            plt.legend(wedges, [f"{label} ({count}, CTR: {ctr:.4f})" for label, count, ctr in zip(labels, counts, ctrs)],
                       title="Categories", loc="center left", bbox_to_anchor=(1, 0, 0.5, 1))

            # 在输出路径中添加筛选条件的值
            output_chart_path = os.path.join(base_path, f'{filter_layers}_{folder}_distribution.png')
            plt.savefig(output_chart_path, bbox_inches='tight')
            plt.close()
            print(f"饼图已保存至: {output_chart_path}")

            # 将图片插入到Excel中
            img = XLImage(output_chart_path)
            sheet.add_image(img, f'H{row}')

            # 写入数据到Excel
            for subfolder, data in stats.items():
                if data['count'] > 0:
                    sheet.cell(row=row, column=1, value=folder)
                    sheet.cell(row=row, column=2, value=subfolder)
                    sheet.cell(row=row, column=3, value=data['count'])
                    sheet.cell(row=row, column=4, value=data['uv'])
                    sheet.cell(row=row, column=5, value=data['click_uv'])
                    sheet.cell(row=row, column=6, value=data['ctr'])
                    row += 1

            row += 15  # 为下一个文件夹的数据和图表留出空间

    # 计算总体统计
    total_stats = {
        'price': {'count': 0, 'uv': 0, 'click_uv': 0},
        'txt': {'count': 0, 'uv': 0, 'click_uv': 0},
        'scene': {'count': 0, 'uv': 0, 'click_uv': 0},
        'white': {'count': 0, 'uv': 0, 'click_uv': 0}
    }

    for folder_stats in all_stats.values():
        for subfolder, data in folder_stats.items():
            total_stats[subfolder]['count'] += data['count']
            total_stats[subfolder]['uv'] += data['uv']
            total_stats[subfolder]['click_uv'] += data['click_uv']

    # 计算总体CTR
    for subfolder in total_stats:
        if total_stats[subfolder]['uv'] > 0:
            total_stats[subfolder]['ctr'] = total_stats[subfolder]['click_uv'] / total_stats[subfolder]['uv']
        else:
            total_stats[subfolder]['ctr'] = 0

    # 打印总体统计结果
    print("\n总体统计结果:")
    for subfolder, data in total_stats.items():
        if data['count'] > 0:
            # print(f"  {subfolder}: 总图片数量 = {data['count']}, 总UV = {data['uv']}, 总Click UV = {data['click_uv']}, 总CTR = {data['ctr']:.4f}")
            continue

    # 绘制总体分布饼图
    total_counts = [data['count'] for data in total_stats.values() if data['count'] > 0]
    total_labels = [subfolder for subfolder, data in total_stats.items() if data['count'] > 0]
    total_ctrs = [data['click_uv'] / data['uv'] if data['uv'] > 0 else 0 for data in total_stats.values() if data['count'] > 0]
    total_colors = [color_map.get(label, 'gray') for label in total_labels]

    if total_counts:
        plt.figure(figsize=(12, 9))
        wedges, texts, autotexts = plt.pie(total_counts, labels=total_labels, autopct='%1.1f%%', startangle=90, colors=total_colors)

        # 添加数量和CTR标签
        for i, (autotext, ctr) in enumerate(zip(autotexts, total_ctrs)):
            autotext.set_text(f'{autotext.get_text()}\n({total_counts[i]})\nCTR: {ctr:.4f}')

        plt.title("Overall Image Distribution")
        plt.axis('equal')

        # 添加图例
        plt.legend(wedges, [f"{label} ({count}, CTR: {ctr:.4f})" for label, count, ctr in zip(total_labels, total_counts, total_ctrs)],
                   title="Categories", loc="center left", bbox_to_anchor=(1, 0, 0.5, 1))

        # output_chart_path = os.path.join(base_path, f'{filter_layers}_overall_distribution.png')
        plt.savefig(output_chart_path_3, bbox_inches='tight')
        plt.close()
        print(f"总体分布饼图已保存至: {output_chart_path_3}")

        # 将总体分布图片插入到Excel中
        img = XLImage(output_chart_path)
        sheet.add_image(img, f'H{row}')

        # 写入总体数据到Excel
        sheet.cell(row=row, column=1, value="Overall")
        for subfolder, data in total_stats.items():
            if data['count'] > 0:
                sheet.cell(row=row, column=2, value=subfolder)
                sheet.cell(row=row, column=3, value=data['count'])
                sheet.cell(row=row, column=4, value=data['uv'])
                sheet.cell(row=row, column=5, value=data['click_uv'])
                sheet.cell(row=row, column=6, value=data['click_uv'] / data['uv'] if data['uv'] > 0 else 0)
                row += 1

    # 在Excel表名中添加筛选条件的值
    sheet.title = f"Image Distribution Stats ({', '.join([str(layer) for layer in filter_layers])})"

    # 保存Excel文件，文件名中添加筛选条件的值
    # excel_output_path = os.path.join(base_path, f'{filter_layers}_image_distribution_stats.xlsx')
    workbook.save(excel_output_path_3)
    print(f"统计结果已保存至Excel文件: {excel_output_path_3}")